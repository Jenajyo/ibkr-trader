import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Today's date formats
today_fmt = datetime.today().strftime("%m-%d-%Y")
today_str = datetime.today().strftime("%Y-%m-%d")

# Paths
base_folder = r"C:/Users/jyoti/Downloads/Stocks"
excel_path = f"{base_folder}/Stocks.xlsx"
volume_csv = f"{base_folder}/all-us-exchanges-price-volume-leaders-{today_fmt}.csv"
top100_csv = f"{base_folder}/top-100-stocks-to-buy-{today_fmt}.csv"

# Load workbook
wb = load_workbook(excel_path, data_only=True)

# Filter logic
def filter_stocks(df):
    # Include rows where:
    # - Medium and Long Term contain 'Buy'
    # - Short Term is either 'Buy' or 'Hold'
    # - Analysts >= 20
    return df[
        (df["Medium Term"].str.contains("Buy", case=False, na=False)) &
        (df["Long Term"].str.contains("Buy", case=False, na=False)) &
        (df["Short Term"].str.contains("Buy|Hold", case=False, na=False)) &
        (pd.to_numeric(df["# Analysts"], errors="coerce") >= 20)
    ]

volume_df = pd.read_csv(volume_csv)
top100_df = pd.read_csv(top100_csv)
filtered_volume_df = filter_stocks(volume_df)
filtered_top100_df = filter_stocks(top100_df)
volume_tickers = set(filtered_volume_df["Symbol"].astype(str).str.upper().str.strip())
top100_tickers = set(filtered_top100_df["Symbol"].astype(str).str.upper().str.strip())

# Extract headers safely
def extract_headers(ws):
    known_headers = {"STOCK", "SYMBOL", "Last", "Buy Price", "Invested", "Current", "Quantity", "Selling Price", "Profit", "Date Added", "DateSold"}
    for row in ws.iter_rows(min_row=1, max_row=10):
        row_values = [str(cell.value).strip() if cell.value else "" for cell in row]
        matches = [h for h in row_values if h in known_headers or h == "Quantiity"]
        if len(matches) >= 5:
            headers = {}
            for idx, val in enumerate(row_values):
                if val == "Quantiity":
                    val = "Quantity"
                headers[val] = idx + 1
            return headers
    raise ValueError("❌ Could not locate a valid header row. Check your Excel file formatting.")

# Normalize ticker symbol strings
def normalize(symbol):
    return str(symbol).strip().upper() if symbol else ""

# Match and print comparison only

def analyze_sheet(ws, filtered_tickers, filtered_df):
    headers = extract_headers(ws)
    symbol_col = headers["SYMBOL"]
    selling_price_col = headers["Selling Price"]

    # Build set of existing unsold tickers
    existing_unsold = set()
    hold_only_symbols = set()
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        symbol_val = row[symbol_col - 1].value
        selling_val = row[selling_price_col - 1].value
        if symbol_val:
            symbol_clean = str(symbol_val).strip().upper()
            if selling_val is None or str(selling_val).strip() == "":
                existing_unsold.add(symbol_clean)

    # Create lookup of short term ratings from filtered data
    filtered_lookup = {str(row["Symbol"]).strip().upper(): row["Short Term"] for _, row in filtered_df.iterrows()}

    # Filter logic based on HOLD status rules
    to_add = []
    to_remove = []
    to_hold = []
    for ticker in filtered_tickers:
        rating = filtered_lookup.get(ticker, "")
        if ticker in existing_unsold:
            if rating.lower() == "hold":
                to_hold.append(ticker)
            else:
                to_hold.append(ticker)
        else:
            if rating.lower() != "hold":
                to_add.append(ticker)

    for ticker in existing_unsold:
        if ticker not in filtered_tickers:
            rating = filtered_lookup.get(ticker, "")
            if rating.lower() != "hold":
                to_remove.append(ticker)

    print("📥 To ADD:", sorted(to_add))
    print("📤 To REMOVE:", sorted(to_remove))
    print("✅ To HOLD:", sorted(to_hold))

        # Print breakdown by ticker
    print("📋 Summary by Ticker:")
    for ticker in sorted(filtered_tickers.union(existing_unsold)):
        short_term = filtered_df.loc[filtered_df["Symbol"].str.upper().str.strip() == ticker, "Short Term"].values
        medium_term = filtered_df.loc[filtered_df["Symbol"].str.upper().str.strip() == ticker, "Medium Term"].values
        long_term = filtered_df.loc[filtered_df["Symbol"].str.upper().str.strip() == ticker, "Long Term"].values

        short_term = short_term[0] if len(short_term) > 0 else "N/A"
        medium_term = medium_term[0] if len(medium_term) > 0 else "N/A"
        long_term = long_term[0] if len(long_term) > 0 else "N/A"

        if ticker in to_add:
            reason = "ADD"
        elif ticker in to_remove:
            reason = "REMOVE"
        elif ticker in to_hold:
            reason = "HOLD"
        else:
            reason = "SKIP"

        print(f"{ticker:<8} | Short: {short_term:<5} | Medium: {medium_term:<5} | Long: {long_term:<5} | Action: {reason}")

    return to_add, to_remove

# Run analysis only
print("📊 TopVolume:")
added_volume, removed_volume = analyze_sheet(wb["TopVolume"], volume_tickers, filtered_volume_df)
print("📊 Top100:")
added_top100, removed_top100 = analyze_sheet(wb["Top100"], top100_tickers, filtered_top100_df)

  # Removed saving step

  # Removed workbook saved message
