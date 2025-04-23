from ib_insync import *
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---- CONFIG ----
excel_file = 'c:/Users/jyoti/Downloads/Orders.xlsx'
target_sheet = 'BUY_Usual'

try:
    # Connect to IBKR
    ib = IB()
    ib.connect('127.0.0.1', 7497, clientId=2)

    # Fetch portfolio with average cost
    positions = ib.portfolio()
    position_dict = {
        (pos.contract.symbol, pos.contract.secType): {
            'quantity': pos.position,
            'avgCost': pos.averageCost
        }
        for pos in positions if pos.contract.secType == 'STK'
    }

    # Read Excel sheet
    sheets = pd.read_excel(excel_file, sheet_name=None)
    if target_sheet not in sheets:
        raise Exception(f"Sheet '{target_sheet}' not found in Excel file.")

    df = sheets[target_sheet]
    df.columns = df.columns.str.strip()

    # Ensure required columns exist
    for col in ['Ticker', 'Quantity', 'Amount']:
        if col not in df.columns:
            df[col] = None

    existing_tickers = set(df['Ticker'].dropna().astype(str).str.strip().tolist())

    # Update existing rows
    for idx, row in df.iterrows():
        ticker = str(row.get('Ticker', '')).strip()
        if ticker:
            key = (ticker, 'STK')
            pos_info = position_dict.get(key, None)
            if pos_info:
                qty = pos_info['quantity']
                avg_cost = pos_info['avgCost']
                amount = math.ceil(qty * avg_cost)

                df.at[idx, 'Quantity'] = qty
                df.at[idx, 'Amount'] = amount

    # Add new holdings
    new_rows = []
    for (symbol, sec_type), data in position_dict.items():
        if symbol not in existing_tickers:
            qty = data['quantity']
            avg_cost = data['avgCost']
            amount = math.ceil(qty * avg_cost)

            new_row = {col: "" for col in df.columns}
            new_row['Ticker'] = symbol
            new_row['Quantity'] = qty
            new_row['Amount'] = amount
            new_rows.append(new_row)

    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    # Write back to Excel
    workbook = load_workbook(excel_file)
    if target_sheet in workbook.sheetnames:
        sheet = workbook[target_sheet]
        sheet.delete_rows(1, sheet.max_row)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

    workbook.save(excel_file)
    print(f"✅ '{target_sheet}' updated with quantity and ceiling-rounded amount.")

    ib.disconnect()

except PermissionError:
    print(f"❌ Permission denied: Please close the file '{excel_file}' and try again.")
except Exception as e:
    print(f"❌ Error: {e}")
